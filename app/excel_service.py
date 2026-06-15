from __future__ import annotations

import io
import os
import re
from pathlib import Path
from typing import BinaryIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side
from openpyxl.styles.borders import Border
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.config import get_settings
from app.ip_utils import ip_in_configured_ranges
from app.matcher import normalize_hostname
from app.models import Machine


HEADERS = [
    "№",
    "Відділ",
    "П І Б",
    None,
    None,
    "Підрозділ",
    "Працює на",
    "Ip",
    "hostname",
    "mac",
    "розетка",
    "інв. ПК",
    "комутатор:порт",
    "Локація",
    "Доступ",
    "примітка",
]

WIDTHS = {
    "A": 5,
    "B": 48,
    "C": 36,
    "D": 3,
    "E": 3,
    "F": 18,
    "G": 16,
    "H": 22,
    "I": 24,
    "J": 18,
    "K": 12,
    "L": 16,
    "M": 22,
    "N": 18,
    "O": 14,
    "P": 18,
}


def _value(value):
    return "" if value is None else value


def _has_wifi_ip(ip_address: str | None) -> bool:
    return ip_in_configured_ranges(ip_address, get_settings().wifi_networks)


def _works_on(machine: Machine) -> str | None:
    if _has_wifi_ip(machine.ip_address):
        return "Ноутбук"
    return machine.works_on


def export_rows(db: Session) -> list[Machine]:
    stmt = (
        select(Machine)
        .where(Machine.is_active.is_(True))
        .order_by(Machine.company, Machine.department, Machine.full_name, Machine.hostname)
    )
    return list(db.scalars(stmt).all())


def workbook_to_bytes(db: Session) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for idx, machine in enumerate(export_rows(db), start=1):
        row = [
            idx,
            machine.position_title,
            machine.full_name,
            None,
            None,
            machine.department or machine.company,
            _works_on(machine),
            machine.ip_address,
            machine.hostname,
            machine.mac_address,
            machine.socket,
            machine.inventory_pc,
            machine.switch_port,
            machine.location,
            machine.oschad_access,
            machine.note,
        ]
        excel_row = idx + 1
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=_value(value))
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for column, width in WIDTHS.items():
        ws.column_dimensions[column].width = width
    ws.column_dimensions["D"].hidden = True
    ws.column_dimensions["E"].hidden = True
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:P{max(ws.max_row, 1)}"

    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


def write_export_file(db: Session, path: str | None = None) -> str:
    settings = get_settings()
    target = Path(path or settings.export_path)
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_bytes(workbook_to_bytes(db))
    return str(target)


def maybe_write_export(db: Session) -> None:
    settings = get_settings()
    if settings.write_export_on_change:
        write_export_file(db, settings.export_path)


def split_cell(value) -> list[str]:
    if value is None:
        return []
    text = str(value).strip()
    if not text:
        return []
    parts = re.split(r"\s*/\s*|[,;\n]+", text)
    return [part.strip() for part in parts if part.strip()]


def _pick(values: list[str], idx: int) -> str | None:
    if not values:
        return None
    if len(values) == 1:
        return values[0]
    return values[idx] if idx < len(values) else None


def import_workbook(file_or_path: str | os.PathLike | BinaryIO, db: Session) -> int:
    wb = load_workbook(file_or_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    imported = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        values = list(row) + [None] * 16
        hostnames = split_cell(values[8])
        if not hostnames:
            continue
        ips = split_cell(values[7])
        for idx, hostname in enumerate(hostnames):
            normalized = normalize_hostname(hostname)
            if not normalized:
                continue
            machine = db.scalar(select(Machine).where(Machine.normalized_hostname == normalized))
            if machine is None:
                machine = Machine(hostname=hostname.strip(), normalized_hostname=normalized)
                db.add(machine)
            machine.hostname = hostname.strip()
            machine.source_type = "virtual" if normalized.startswith("vm-") else (machine.source_type or "physical")
            machine.position_title = values[1] or machine.position_title
            machine.full_name = values[2] or machine.full_name
            machine.company = values[5] or machine.company
            machine.works_on = values[6] or machine.works_on
            machine.ip_address = _pick(ips, idx) or machine.ip_address
            if _has_wifi_ip(machine.ip_address):
                machine.works_on = "Ноутбук"
            machine.mac_address = values[9] or machine.mac_address
            machine.socket = values[10] or machine.socket
            machine.inventory_pc = values[11] or machine.inventory_pc
            machine.switch_port = values[12] or machine.switch_port
            machine.location = values[13] or machine.location
            machine.oschad_access = values[14] or machine.oschad_access
            machine.note = values[15] or machine.note
            if machine.match_status == "not_synced":
                machine.match_status = "imported"
            imported += 1
    db.commit()
    maybe_write_export(db)
    return imported
